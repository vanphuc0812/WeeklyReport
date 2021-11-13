import openpyxl
from openpyxl.workbook import workbook
from openpyxl.worksheet import worksheet
import sys

print('----------------------This script is created by VanPhuc--------------------------------')
print('----------------------This script is created by VanPhuc--------------------------------')
print('----------------------This script is created by VanPhuc--------------------------------')
print('')
week =44 #str(sys.argv[1])
line_of_week =94 #int(sys.argv[2])
print('Preparing environment to create report for week ' + str(week) + ' in line '+ str(line_of_week))

#Sample report
try:
    REPORT = openpyxl.load_workbook('./SAMPLE REPORT/Sample_Report.xlsx')
    REPORT_ws = REPORT.active
except: 
    print("    File Sample_Report.xlsx not exist")


#taiwan
try: 
    TAIWAN_LASTWEEK = openpyxl.load_workbook('./TAIWAN/'+str(int(week)-1)+'/TAIWAN.xlsx')['Sheet1']
except: 
    print("    File TAIWAN/LASTWEEK.xlsx not exist")
try: 
    TAIWAN = openpyxl.load_workbook('./TAIWAN/'+str(week)+'/TAIWAN.xlsx')['Sheet1']
except: 
    print("    File TAIWAN.xlsx not exist")

#japan
try: 
    JAPAN_LASTWEEK = openpyxl.load_workbook('./JAPAN/'+str(int(week)-1)+'/JAPAN.xlsx')['Sheet1']
except: 
    print("    File JAPAN/LASTWEEK.xlsx not exist")
try:
    JAPAN  = openpyxl.load_workbook('./JAPAN/'+str(week)+'/JAPAN.xlsx')['Sheet1']
except: 
    print("    File JAPAN.xlsx not exist")

#med
try: 
    MED_LASTWEEK = openpyxl.load_workbook('./MED/'+str(int(week)-1)+'/MED.xlsx')['Sheet1']
except: 
    print("    File MED/LASTWEEK.xlsx not exist")
try:
    MED = openpyxl.load_workbook('./MED/'+str(week)+'/MED.xlsx')['Sheet1']
except: 
    print("    File MED.xlsx not exist")
#euro
try: 
    EURO_LASTWEEK = openpyxl.load_workbook('./EURO/'+str(int(week)-1)+'/EURO.xlsx')['Sheet1']
except: 
    print("    File EURO/LASTWEEK.xlsx not exist")
try:
    EURO  = openpyxl.load_workbook('./EURO/'+str(week)+'/EURO.xlsx')['Sheet1']
except: 
    print("    File EURO.xlsx not exist")

def find_indexOf(key,worksheet):
    ro = 0
    col = 0
    for i in range(1,worksheet.max_row+1):
        for j in range(1,worksheet.max_column+1):
            if worksheet.cell(i,j).value != None:
                if str(worksheet.cell(i,j).value).strip() == key:
                    ro=i
                    col=j
    return ro,col

def find_week(week) :
    col=0
    for i in range (50,10000):
        if REPORT_ws._get_cell(row=line_of_week, column=i).value == int(week):
            col=i
            break
    return col

def get_colunm_total(key, worksheet): 
    total=0
    row,col=find_indexOf(key, worksheet)
    for i in range(row+1,worksheet.max_row):
        if  worksheet.cell(i,col).value != 0 :
            total += float(worksheet.cell(i,col).value)
    return total
def get_column_num(key, worksheet): 
    num=0
    row,col=find_indexOf(key, worksheet)
    for i in range(row+1,worksheet.max_row):
        if  worksheet.cell(i,col).value != 0 :
            num += 1
    return num

def get_list_shipcode(worksheet1,worksheet2) :
    result= []
    row1,col1=find_indexOf("Shipper Code", worksheet1)
    row2,col2=find_indexOf("Shipper Code", worksheet2)
    for i in range(row1+1,worksheet1.max_row):
        result.append( worksheet1.cell(i,col1).value)
    for i in range(row2+1,worksheet2.max_row):
        result.append( worksheet2.cell(i,col2).value)
    result=list(set(result))
    return result
def get_tues(worksheet,shipcode):
    tues=0
    row,col= find_indexOf(shipcode,worksheet)
    if row != 0 :
        row1,col1= find_indexOf('FCL/20\'',worksheet)
        row2,col2= find_indexOf('FCL/40\'',worksheet)
        row3,col3= find_indexOf('FCL/40\'HQ',worksheet)
        row4,col4= find_indexOf('FCL/45\'',worksheet)
        tues = int(worksheet.cell(row,col1).value) + int(worksheet.cell(row,col2).value)*2 +int(worksheet.cell(row,col3).value)*2+int(worksheet.cell(row,col4).value)*2
    return tues
def compare_teus(worksheet1,worksheet2,list):
    list_result=[]
    for code in list:
        list_result.append(get_tues(worksheet1,code) - get_tues(worksheet2,code))
    return list_result
def sort(list1, list2):
    swapped = True
    while swapped:
        swapped = False
        for i in range(len(list1) - 1):
            if list1[i] > list1[i + 1]:
                # Swap the elements
                list1[i], list1[i + 1] = list1[i + 1], list1[i]
                list2[i], list2[i + 1] = list2[i + 1], list2[i]
                # Set the flag to True so we'll loop again
                swapped = True

def check_exist(shipcode,worksheet):
    result=False
    row,col= find_indexOf('Shipper Code',worksheet)
    for i in range(row +1, worksheet.max_row):
            if worksheet.cell(i,col).value == shipcode: result=True
    return result
#--------------------------------------------------------------------------------------------------------------
def main():
    base_col=find_week(week)
    print(find_indexOf("GND TOTAL",TAIWAN))
    print(get_colunm_total("HB/L SET(s)",TAIWAN))
    if(base_col!=0):
        print('')
        print('-------------------------------Data for TAIWAN-------------------------------')
        try:
            shipcode_list=get_list_shipcode(TAIWAN,TAIWAN_LASTWEEK)
            list_change=compare_teus(TAIWAN,TAIWAN_LASTWEEK,shipcode_list)
            sort(list_change,shipcode_list)
            print(shipcode_list)
            print(list_change)
        except:
            print('')
        print('')
        print('-------------------------------Data for JAPAN-------------------------------')
        try:
            JP_shipcode_list=get_list_shipcode(JAPAN,JAPAN_LASTWEEK)
            JP_list_change=compare_teus(JAPAN,JAPAN_LASTWEEK,JP_shipcode_list)
            sort(JP_list_change,JP_shipcode_list)
            print(JP_shipcode_list)
            print(JP_list_change)
        except:
            print('')
        print('')
        print('-------------------------------Data for EURO-------------------------------')
        try:
            EU_shipcode_list=get_list_shipcode(EURO,EURO_LASTWEEK)
            EU_list_change=compare_teus(EURO,EURO_LASTWEEK,EU_shipcode_list)
            sort(EU_list_change,EU_shipcode_list)
            print(EU_shipcode_list)
            print(EU_list_change)
        except:
            print('')
        print('')
        print('-------------------------------Data for MED-------------------------------')
        try:
            MED_shipcode_list=get_list_shipcode(MED,MED_LASTWEEK)
            MED_list_change=compare_teus(MED,MED_LASTWEEK,MED_shipcode_list)
            sort(MED_list_change,MED_shipcode_list)
            print(MED_shipcode_list)
            print(MED_list_change)
        except:
            print('')
        print('')

        #TAIWAN
        try:
            #shpt/LCL
            REPORT_ws._get_cell(row=int(line_of_week)+5,column=base_col).value= get_column_num('CFS(cbm)',TAIWAN)
            #CBM/LCL
            REPORT_ws._get_cell(row=int(line_of_week)+5,column=base_col+2).value = get_colunm_total('CFS(cbm)',TAIWAN)
            #consol 
            REPORT_ws._get_cell(row=int(line_of_week)+6,column=base_col+2).value = get_colunm_total('CNSL20(cbm)',TAIWAN)+ get_colunm_total('CNSL40(cbm)',TAIWAN)
            REPORT_ws._get_cell(row=int(line_of_week)+6,column=base_col).value = get_column_num('CNSL20(cbm)',TAIWAN)+ get_column_num('CNSL40(cbm)',TAIWAN)
            #Teus/FCL
            REPORT_ws._get_cell(row=int(line_of_week)+4,column=base_col+1).value = get_colunm_total('FCL/20\'',TAIWAN)+get_colunm_total('FCL/40\'',TAIWAN)*2+get_colunm_total('FCL/40\'HQ',TAIWAN)*2+get_colunm_total('FCL/45\'',TAIWAN)*2
            #shpt/FCL
            REPORT_ws._get_cell(row=int(line_of_week)+4,column=base_col).value = get_colunm_total('HB/L SET(s)',TAIWAN) - (get_column_num('CNSL20(cbm)',TAIWAN)+ get_column_num('CNSL40(cbm)',TAIWAN) + get_column_num('CFS(cbm)',TAIWAN) )
            print('    Successfully finish report for TAIWAN')
        except:
            print('    Can not finish report for TAIWAN')
    #JAPAN
        try:
            #shpt/LCL
            REPORT_ws._get_cell(row=int(line_of_week)+5,column=base_col+3).value = get_column_num('CFS(cbm)',JAPAN)
            #CBM/LCL
            REPORT_ws._get_cell(row=int(line_of_week)+5,column=base_col+5).value = get_colunm_total('CFS(cbm)',JAPAN)
            #consol 
            REPORT_ws._get_cell(row=int(line_of_week)+6,column=base_col+5).value = get_colunm_total('CNSL20(cbm)',JAPAN)+ get_colunm_total('CNSL40(cbm)',JAPAN)
            REPORT_ws._get_cell(row=int(line_of_week)+6,column=base_col+3).value = get_column_num('CNSL20(cbm)',JAPAN)+ get_column_num('CNSL40(cbm)',JAPAN)
            #Teus/FCL
            REPORT_ws._get_cell(row=int(line_of_week)+4,column=base_col+4).value = get_colunm_total('FCL/20\'',JAPAN)+get_colunm_total('FCL/40\'',JAPAN)*2+get_colunm_total('FCL/40\'HQ',JAPAN)*2+get_colunm_total('FCL/45\'',JAPAN)*2
            #shpt/FCL
            REPORT_ws._get_cell(row=int(line_of_week)+4,column=base_col+3).value = get_colunm_total('HB/L SET(s)',JAPAN) - (get_column_num('CNSL20(cbm)',JAPAN)+ get_column_num('CNSL40(cbm)',JAPAN) + get_column_num('CFS(cbm)',JAPAN) )
            print('    Successfully finish report for JAPAN')
        except:
            print('    Can not finish report for JAPAN')
    #EURO
        try:
            #shpt/LCL
            REPORT_ws._get_cell(row=int(line_of_week)+16,column=base_col+6).value = get_column_num('CFS(cbm)',EURO)
            #CBM/LCL
            REPORT_ws._get_cell(row=int(line_of_week)+16,column=base_col+8).value = get_colunm_total('CFS(cbm)',EURO)
            #consol 
            REPORT_ws._get_cell(row=int(line_of_week)+17,column=base_col+8).value = get_colunm_total('CNSL20(cbm)',EURO)+ get_colunm_total('CNSL40(cbm)',EURO)
            REPORT_ws._get_cell(row=int(line_of_week)+17,column=base_col+6).value = get_column_num('CNSL20(cbm)',EURO)+ get_column_num('CNSL40(cbm)',EURO)
            #Teus/FCL
            REPORT_ws._get_cell(row=int(line_of_week)+15,column=base_col+7).value = get_colunm_total('FCL/20\'',EURO)+get_colunm_total('FCL/40\'',EURO)*2+get_colunm_total('FCL/40\'HQ',EURO)*2+get_colunm_total('FCL/45\'',EURO)*2
            #shpt/FCL
            REPORT_ws._get_cell(row=int(line_of_week)+15,column=base_col+6).value = get_colunm_total('HB/L SET(s)',EURO) - (get_column_num('CNSL20(cbm)',EURO)+ get_column_num('CNSL40(cbm)',EURO) + get_column_num('CFS(cbm)',EURO) )
            print('    Successfully finish report for EURO')
        except:
            print('    Can not finish report for EURO')
    #MED
        try:
            #shpt/LCL
            REPORT_ws._get_cell(row=int(line_of_week)+16,column=base_col+9).value = get_column_num('CFS(cbm)',MED)
            #CBM/LCL
            REPORT_ws._get_cell(row=int(line_of_week)+16,column=base_col+11).value = get_colunm_total('CFS(cbm)',MED)
            #consol 
            REPORT_ws._get_cell(row=int(line_of_week)+17,column=base_col+11).value = get_colunm_total('CNSL20(cbm)',MED)+ get_colunm_total('CNSL40(cbm)',MED)
            REPORT_ws._get_cell(row=int(line_of_week)+17,column=base_col+9).value = get_column_num('CNSL20(cbm)',MED)+ get_column_num('CNSL40(cbm)',MED)
            #Teus/FCL
            REPORT_ws._get_cell(row=int(line_of_week)+15,column=base_col+10).value = get_colunm_total('FCL/20\'',MED)+get_colunm_total('FCL/40\'',MED)*2+get_colunm_total('FCL/40\'HQ',MED)*2+get_colunm_total('FCL/45\'',MED)*2
            #shpt/FCL
            REPORT_ws._get_cell(row=int(line_of_week)+15,column=base_col+9).value = get_colunm_total('HB/L SET(s)',MED) - (get_column_num('CNSL20(cbm)',MED)+ get_column_num('CNSL40(cbm)',MED) + get_column_num('CFS(cbm)',MED) )
            print('    Successfully finish report for MED')
        except:
            print('    Can not finish report for MED')
    #CMT FOR TAIWAN
        try:
            counter=0
            cmt1=''
            cmt2=''
            if get_colunm_total('FCL/20\'',TAIWAN_LASTWEEK) +get_colunm_total('FCL/40\'',TAIWAN_LASTWEEK)*2 +get_colunm_total('FCL/40\'HQ',TAIWAN_LASTWEEK)*2 > get_colunm_total('FCL/20\'',TAIWAN) +get_colunm_total('FCL/40\'',TAIWAN)*2 +get_colunm_total('FCL/40\'HQ',TAIWAN)*2 :
                REPORT_ws._get_cell(row=26,column=base_col).value = 'TAIWAN: volume of this week had been decreased than last week due to:'
                for i in range(len(list_change) - 1):
                    if(counter<10):
                        counter+=1
                        if check_exist(shipcode_list[i],TAIWAN):
                            if list_change[i] <0:
                                cmt1+= str(shipcode_list[i]) + ' decreased ' + str(abs(list_change[i])) + ' teus, '
                        else:   cmt2+= str(shipcode_list[i]) + ' , '
                REPORT_ws._get_cell(row=27,column=base_col).value =cmt1 
                REPORT_ws._get_cell(row=28,column=base_col).value ='No cargo: '+cmt2 
            elif get_colunm_total('FCL/20\'',TAIWAN_LASTWEEK) +get_colunm_total('FCL/40\'',TAIWAN_LASTWEEK)*2 +get_colunm_total('FCL/40\'HQ',TAIWAN_LASTWEEK)*2 < get_colunm_total('FCL/20\'',TAIWAN) +get_colunm_total('FCL/40\'',TAIWAN)*2 +get_colunm_total('FCL/40\'HQ',TAIWAN)*2 :
                REPORT_ws._get_cell(row=26,column=base_col).value = 'TAIWAN: volume of this week had been increased than last week due to:'
                if len(list_change) <10 :
                    for i in range(len(list_change)):
                        if check_exist(shipcode_list[i],TAIWAN):
                            if list_change[i] >0:
                                cmt1+= str(shipcode_list[i]) + ' increased ' + str(list_change[i]) + ' teus, ' 
                        else:   cmt2+= str(shipcode_list[i])
                    REPORT_ws._get_cell(row=27,column=base_col).value =cmt1 
                    REPORT_ws._get_cell(row=28,column=base_col).value ='No cargo: ' + cmt2 
                else:
                    for i in range(len(list_change)-10 ,len(list_change)):
                        if check_exist(shipcode_list[i],TAIWAN):
                            if list_change[i] >0:
                                cmt1+= str(shipcode_list[i]) + ' increased ' + str(list_change[i]) + ' teus, ' 
                        else:   cmt2+= str(shipcode_list[i])
                    REPORT_ws._get_cell(row=27,column=base_col).value =cmt1 
                    REPORT_ws._get_cell(row=28,column=base_col).value ='No cargo: ' + cmt2 
            else: 
                REPORT_ws._get_cell(row=26,column=base_col).value = 'TAIWAN: volume of this week is no change.'
        except:
            print('    Can not find data to make comment for TAIWAN')

        #cmt for japan
        try:
            counter=0
            cmt1=''
            cmt2=''
            if get_colunm_total('FCL/20\'',JAPAN_LASTWEEK) +get_colunm_total('FCL/40\'',JAPAN_LASTWEEK)*2 +get_colunm_total('FCL/40\'HQ',JAPAN_LASTWEEK)*2 > get_colunm_total('FCL/20\'',JAPAN) +get_colunm_total('FCL/40\'',JAPAN)*2 +get_colunm_total('FCL/40\'HQ',JAPAN)*2 :
                REPORT_ws._get_cell(row=22,column=base_col).value = 'JAPAN: volume of this week had been decreased than last week due to:'
                for i in range(len(JP_list_change) - 1):
                    if(counter<10):
                        counter+=1
                        if check_exist(JP_shipcode_list[i],JAPAN):
                            if JP_list_change[i]<0:
                                cmt1+= str(JP_shipcode_list[i]) + ' decreased ' + str(abs(JP_list_change[i])) + ' teus, '
                        else:   cmt2+= str(JP_shipcode_list[i]) + ' , '
                REPORT_ws._get_cell(row=23,column=base_col).value =cmt1 
                REPORT_ws._get_cell(row=24,column=base_col).value ='No cargo: '+cmt2 
            elif get_colunm_total('FCL/20\'',JAPAN_LASTWEEK) +get_colunm_total('FCL/40\'',JAPAN_LASTWEEK)*2 +get_colunm_total('FCL/40\'HQ',JAPAN_LASTWEEK)*2 < get_colunm_total('FCL/20\'',JAPAN) +get_colunm_total('FCL/40\'',JAPAN)*2 +get_colunm_total('FCL/40\'HQ',JAPAN)*2 :
                REPORT_ws._get_cell(row=22,column=base_col).value = 'JAPAN: volume of this week had been increased than last week due to:'
                if len(JP_list_change) <10 :
                    for i in range(len(JP_list_change)):
                        if check_exist(JP_shipcode_list[i],JAPAN):
                            if JP_list_change[i]>0:
                                cmt1+= str(JP_shipcode_list[i]) + ' increased ' + str(JP_list_change[i]) + ' teus, ' 
                        else:   cmt2+= str(JP_shipcode_list[i])
                    REPORT_ws._get_cell(row=23,column=base_col).value =cmt1 
                    REPORT_ws._get_cell(row=24,column=base_col).value ='No cargo: ' + cmt2 
                else:
                    for i in range(len(JP_list_change)-10 ,len(JP_list_change)):
                        if check_exist(JP_shipcode_list[i],JAPAN):
                            if JP_list_change[i]>0:
                                cmt1+= str(JP_shipcode_list[i]) + ' increased ' + str(JP_list_change[i]) + ' teus, ' 
                        else:   cmt2+= str(JP_shipcode_list[i])
                    REPORT_ws._get_cell(row=23,column=base_col).value =cmt1 
                    REPORT_ws._get_cell(row=24,column=base_col).value ='No cargo: ' + cmt2 
            else: 
                REPORT_ws._get_cell(row=22,column=base_col).value = 'JAPAN: volume of this week is no change.'
        except:
            print('    Can not find data to make comment for JAPAN')
#CMT FOR EURO
        try:
            counter=0
            cmt1=''
            cmt2=''
            if get_colunm_total('FCL/20\'',EURO_LASTWEEK) +get_colunm_total('FCL/40\'',EURO_LASTWEEK)*2 +get_colunm_total('FCL/40\'HQ',EURO_LASTWEEK)*2 > get_colunm_total('FCL/20\'',EURO) +get_colunm_total('FCL/40\'',EURO)*2 +get_colunm_total('FCL/40\'HQ',EURO)*2 :
                REPORT_ws._get_cell(row=30,column=base_col).value = 'EURO: volume of this week had been decreased than last week due to:'
                for i in range(len(EU_list_change) - 1):
                    if(counter<10):
                        counter+=1
                        if check_exist(EU_shipcode_list[i],EURO):
                            if EU_list_change[i]<0:
                                cmt1+= str(EU_shipcode_list[i]) + ' decreased ' + str(abs(EU_list_change[i])) + ' teus, '
                        else:   cmt2+= str(EU_shipcode_list[i]) + ' , '
                REPORT_ws._get_cell(row=31,column=base_col).value =cmt1 
                REPORT_ws._get_cell(row=32,column=base_col).value ='No cargo: '+cmt2 
            elif get_colunm_total('FCL/20\'',EURO_LASTWEEK) +get_colunm_total('FCL/40\'',EURO_LASTWEEK)*2 +get_colunm_total('FCL/40\'HQ',EURO_LASTWEEK)*2 < get_colunm_total('FCL/20\'',EURO) +get_colunm_total('FCL/40\'',EURO)*2 +get_colunm_total('FCL/40\'HQ',EURO)*2 :
                REPORT_ws._get_cell(row=30,column=base_col).value = 'EURO: volume of this week had been increased than last week due to:'
                if len(EU_list_change) <10 :
                    for i in range(len(EU_list_change)):
                        if check_exist(EU_shipcode_list[i],EURO):
                            if EU_list_change[i]>0:
                                cmt1+= str(EU_shipcode_list[i]) + ' increased ' + str(EU_list_change[i]) + ' teus, ' 
                        else:   cmt2+= str(EU_shipcode_list[i])
                    REPORT_ws._get_cell(row=31,column=base_col).value =cmt1 
                    REPORT_ws._get_cell(row=32,column=base_col).value ='No cargo: ' + cmt2 
                else :
                    for i in range(len(EU_list_change)-10,len(EU_list_change)):
                        if check_exist(EU_shipcode_list[i],EURO):
                            if EU_list_change[i]>0:
                                cmt1+= str(EU_shipcode_list[i]) + ' increased ' + str(EU_list_change[i]) + ' teus, ' 
                        else:   cmt2+= str(EU_shipcode_list[i])
                    REPORT_ws._get_cell(row=31,column=base_col).value =cmt1 
                    REPORT_ws._get_cell(row=32,column=base_col).value ='No cargo: ' + cmt2 
            else: 
                REPORT_ws._get_cell(row=30,column=base_col).value = 'EURO: volume of this week is no change.'
        except:
            print('    Can not find data to make comment for EURO')
#CMT FOR MED
        try:
            counter=0
            cmt1=''
            cmt2=''
            if get_colunm_total('FCL/20\'',MED_LASTWEEK) +get_colunm_total('FCL/40\'',MED_LASTWEEK)*2 +get_colunm_total('FCL/40\'HQ',MED_LASTWEEK)*2 > get_colunm_total('FCL/20\'',MED) +get_colunm_total('FCL/40\'',MED)*2 +get_colunm_total('FCL/40\'HQ',MED)*2 :
                    REPORT_ws._get_cell(row=33,column=base_col).value = 'MED: volume of this week had been decreased than last week due to:'
                    for i in range(len(MED_list_change) - 1):
                        if(counter<10):
                            counter+=1
                            if check_exist(MED_shipcode_list[i],MED):
                                if MED_list_change[i]<0:
                                    cmt1+= str(MED_shipcode_list[i]) + ' decreased ' + str(abs(MED_list_change[i])) + ' teus, '
                            else:   cmt2+= str(MED_shipcode_list[i]) + ' , '
                    REPORT_ws._get_cell(row=34,column=base_col).value =cmt1 
                    REPORT_ws._get_cell(row=35,column=base_col).value ='No cargo: '+cmt2 
            elif get_colunm_total('FCL/20\'',MED_LASTWEEK) +get_colunm_total('FCL/40\'',MED_LASTWEEK)*2 +get_colunm_total('FCL/40\'HQ',MED_LASTWEEK)*2 < get_colunm_total('FCL/20\'',MED) +get_colunm_total('FCL/40\'',MED)*2 +get_colunm_total('FCL/40\'HQ',MED)*2 :
                    REPORT_ws._get_cell(row=33,column=base_col).value = 'MED: volume of this week had been increased than last week due to:'
                    if len(MED_list_change) <10 :
                        for i in range(len(MED_list_change)):
                            if check_exist(MED_shipcode_list[i],MED):
                                if MED_list_change[i]>0:
                                    cmt1+= str(MED_shipcode_list[i]) + ' increased ' + str(MED_list_change[i]) + ' teus, ' 
                            else:   cmt2+= str(MED_shipcode_list[i])
                        REPORT_ws._get_cell(row=34,column=base_col).value =cmt1 
                        REPORT_ws._get_cell(row=35,column=base_col).value ='No cargo: ' + cmt2
                    else:
                        for i in range(len(MED_list_change)-10 ,len(MED_list_change)):
                            if check_exist(MED_shipcode_list[i],MED):
                                if MED_list_change[i]>0:
                                    cmt1+= str(MED_shipcode_list[i]) + ' increased ' + str(MED_list_change[i]) + ' teus, ' 
                            else:   cmt2+= str(MED_shipcode_list[i])
                        REPORT_ws._get_cell(row=34,column=base_col).value =cmt1 
                        REPORT_ws._get_cell(row=35,column=base_col).value ='No cargo: ' + cmt2 
            else: 
                    REPORT_ws._get_cell(row=33,column=base_col).value = 'MED: volume of this week is no change.'
        except:
            print('    Can not find data to make comment for MED')

        REPORT.save('./Report.xlsx')
        print('.......................................................................................')
        print('')
        print('Suscessfully create report')
    else:
        print('The week you choose is not exist in the sample report')



    print('')
    print('----------------------This script is created by VanPhuc--------------------------------')
    print('----------------------This script is created by VanPhuc--------------------------------')
    print('----------------------This script is created by VanPhuc--------------------------------')
    print('')
if __name__=="__main__":
    main()
